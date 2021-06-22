import openpyxl


class HandleExcel:
    """用来操作excel文件的类"""

    def __init__(self, filename, sheetname):
        """
        初始化对象属性
        :param filename: excel文件路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """读取excel中的数据"""
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheetname]
        # 按行获取所有的数据，转换为列表
        rows_data = list(sh.rows)
        # 创建一个空列表用来保存所有的用例数据
        cases_data = []
        # 获取表单中的表头数据，放入title这个列表中
        title = []
        for i in rows_data[0]:
            title.append(i.value)
        # 获取除表头之外的其他行数据
        for item in rows_data[1:]:
            # 每遍历出来一行数据，就创建一个空列表，来存放该行数据
            values = []
            for i in item:
                values.append(i.value)
            # 将该行的数据和表头进行打包，转换为字典
            case = dict(zip(title, values))
            # 将该行数据打包的字典，放入cases_data中
            cases_data.append(case)
        # 返回读取出来的所有数据
        return cases_data

    def write_data(self, row, column, value):
        """
        写入数据
        :param row: 行
        :param column: 列表
        :param value: 写入的值
        :return:
        """
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheetname]
        # 根据行、列去写入内容
        sh.cell(row=row, column=column, value=value)
        # 把工作簿保存为文件
        wb.save(self.filename)
