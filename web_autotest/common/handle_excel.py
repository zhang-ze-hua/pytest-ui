import openpyxl


class HandleExcel:
    """用来操作excel文件的类"""

    def __init__(self, filename, sheetname):
        """
        初始化对象属性
        :param filename: excel文件路径
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """将excel中的数据组装成用例数据返回"""
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheetname]
        # 按行获取所有的数据，转换为列表
        rows_data = list(sh.rows)
        # 创建一个空列表用来保存所有的用例数据
        cases_data = []
        # 获取表单中的表头数据，放入title这个列表中
        title = []
        for i in rows_data[0]:
            title.append(i.value)
        # result 的列数
        index = title.index("result")
        column = index + 1
        # 获取除表头之外的其他行数据
        for item in rows_data[1:]:
            # 每遍历出来一行数据，就创建一个空列表，来存放该行数据
            values = []
            for i in item:
                values.append(i.value)
            # 将该行的数据和表头进行打包，转换为字典
            case = dict(zip(title, values))
            # 将该行数据打包的字典，放入cases_data中
            cases_data.append(case)
        # 返回用例数据，result列数
        return cases_data, column

    def write_data(self, row, column, value):
        """
        写入数据
        :param row: 行
        :param column: 列表
        :param value: 写入的值
        :return:
        """
        # 获取工作簿对象
        wb = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = wb[self.sheetname]
        # 根据行、列去写入内容
        sh.cell(row=row, column=column, value=value)
        # 把工作簿保存为文件
        wb.save(self.filename)

    def case_classify(self):
        """用例分类，正例、反例"""
        success_case_data = []
        error_alert_data = []
        error_case_data = []
        cases_data, column = self.read_data()
        print(cases_data)
        print(type(cases_data))
        for case_data in cases_data:
            print(case_data)
            print(type(case_data))
            if case_data["type"] == "正例":
                success_case_data.append(case_data)
            elif case_data["type"] == "反例-页面提示":
                error_case_data.append(case_data)
            elif case_data["type"] == "反例-弹窗提示":
                error_alert_data.append(case_data)
        return success_case_data, error_case_data, error_alert_data, column
